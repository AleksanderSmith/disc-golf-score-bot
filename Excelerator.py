from openpyxl import Workbook, load_workbook
from Score import Score

class Excelerator():
  
  def open_workbook(self):
    self.workbook = load_workbook(filename=self.filename)
    
  def create_workbook(self):
    self.workbook = Workbook()
  
  

  def save_workbook(self):
    self.workbook.save(self.filename)
    
  def insert_score(self, data):
    pass
  